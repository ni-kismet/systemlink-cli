"""Generate a sample spec template Excel file for the spec compliance notebooks."""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def create_spec_template() -> None:
    """Create a comprehensive spec template Excel file with documentation and examples."""
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SpecTemplate"

    # Define fills for color legend
    editable_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    column_type_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    std_column_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    description_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # Title
    ws["A3"] = "Specification Template File Notes"
    ws["A3"].font = Font(bold=True, size=12)

    current_row = 4

    # File Notes (1-9)
    file_notes = [
        (
            "1. The spreadsheet tab names in the file you upload to NI Specification "
            "Compliance Manager must be identical to the spreadsheet tab names in the "
            'template file: One sheet named "SpecTemplate" and number of sheets with names '
            'starting with "Spec_" (Name of the sheet should not be "Spec_Data")'
        ),
        (
            "2. The SpecTemplate spreadsheet tab contains only reference material. Data on "
            "this tab does not appear in the NI Specification Compliance Manager user "
            "interface."
        ),
        (
            "3. The Spec_<Your Category 1> & Spec_<Your Category 2> spreadsheet tabs "
            "contains a baseline template to use as a starting point for your product "
            'specifications of type "Parametric" and "Functional" respectively. '
            "Information on the Spec_<Your Category 1> & Spec_<Your Category 2> tabs will "
            "appear in the NI Specification Compliance Manager user interface."
        ),
        (
            "4. A specification file can have both Parametric and Functional specifications "
            "or either one of them."
        ),
        (
            "5. Refer to the format guidelines for each type of column to ensure your "
            "product specifications align with the requirements for upload to NI "
            "Specification Compliance Manager."
        ),
        (
            "6. Refer to the tables starting in column T of the SpecTemplate tab for "
            "examples of correct formatting for the Spec_<Your Category> tabs."
        ),
        (
            "7. The total number of unique (in terms of Column Name) STD, COND & INF "
            'columns from all sheets, with names starting with "Spec_" combined, should not '
            "exceed 1000."
        ),
        (
            "8. New sheets named 'Spec_<Your Category>' can be added as per the requirement. "
            "Unused 'Spec_<Your Category>' sheets should be removed from the file before "
            "uploading to the NI Spec Compliance Manager application."
        ),
        (
            "9. NI Specification Compliance Manager application supports upload of Spec file "
            "in former format supporting only Parametric specs (File with 'SpecTemplate' "
            "and 'Spec_Data' tabs only)"
        ),
    ]

    for note in file_notes:
        ws[f"A{current_row}"] = note
        ws[f"A{current_row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    current_row += 1

    # Template Color Legend
    ws[f"A{current_row}"] = "Template Color Legend"
    ws[f"A{current_row}"].font = Font(bold=True, size=11)
    current_row += 1

    ws[f"A{current_row}"] = "Fill Color"
    ws[f"B{current_row}"] = "Description"
    ws[f"A{current_row}"].font = Font(bold=True)
    ws[f"B{current_row}"].font = Font(bold=True)
    ws[f"A{current_row}"].fill = description_fill
    ws[f"B{current_row}"].fill = description_fill
    current_row += 1

    color_legend = [
        (
            editable_fill,
            "Editable cell that you can change as needed for your product specifications.",
        ),
        (
            column_type_fill,
            "Defines the column type, which determines the formatting rules that apply and how NI Specification Compliance Manager processes the information in the column during the file upload process.",
        ),
        (std_column_fill, "Column names for standard (STD) columns. Do not edit these cells."),
        (description_fill, "Description of the row. Do not edit these cells."),
    ]

    for fill, desc in color_legend:
        ws[f"A{current_row}"].fill = fill
        ws[f"B{current_row}"] = desc
        ws[f"B{current_row}"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    current_row += 1

    # Meta Data Overview
    ws[f"A{current_row}"] = "Meta Data Overview"
    ws[f"A{current_row}"].font = Font(bold=True, size=11)
    current_row += 1

    ws[f"A{current_row}"] = "Meta Data Name"
    ws[f"B{current_row}"] = "Description"
    ws[f"C{current_row}"] = "Entry Required?"
    ws[f"D{current_row}"] = "Entry Format Requirements"
    for col in ["A", "B", "C", "D"]:
        ws[f"{col}{current_row}"].font = Font(bold=True)
        ws[f"{col}{current_row}"].fill = description_fill
    current_row += 1

    meta_data = [
        (
            "Category Name",
            "Category to classify the product specification. NI Specification Compliance Manager will use category information to group product specs under the specified category in the Web App.",
            "Yes",
            "Unique entries per sheet.",
        ),
        (
            "Type",
            "Type of the specifications present under a category",
            "Yes",
            'Should be either "Parametric" or "Functional". Any other value is not allowed.',
        ),
    ]

    for name, desc, required, format_req in meta_data:
        ws[f"A{current_row}"] = name
        ws[f"B{current_row}"] = desc
        ws[f"C{current_row}"] = required
        ws[f"D{current_row}"] = format_req
        for col in ["B", "D"]:
            ws[f"{col}{current_row}"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    current_row += 1

    # Column Type Overview
    ws[f"A{current_row}"] = "Column Type Overview"
    ws[f"A{current_row}"].font = Font(bold=True, size=11)
    current_row += 1

    ws[f"A{current_row}"] = "Column Type"
    ws[f"B{current_row}"] = "Abbreviation"
    ws[f"C{current_row}"] = "Description"
    for col in ["A", "B", "C"]:
        ws[f"{col}{current_row}"].font = Font(bold=True)
        ws[f"{col}{current_row}"].fill = description_fill
    current_row += 1

    column_types = [
        (
            "Standard",
            "STD",
            'A required column for product specifications in NI Specification Compliance Manager. Do not change standard column names. The set of standard column names that should be present varies for "Parametric" and "Functional" type specifications. Do not insert columns between standard columns. All standard columns must be present in the specification file.',
        ),
        (
            "Condition",
            "COND",
            'An optional column with a customizable name. Condition column names can contain a maximum of 32 characters. Specifying a unit of measure for condition columns is optional. Bracket characters "[" and "]" are not allowed in condition column names. The specification file can contain up to 20 condition columns. Same rules mentioned here applies for both "Parametric" and "Functional" type specifications.',
        ),
        (
            "Information",
            "INF",
            'An optional column with a customizable name. Information column names can contain a maximum of 32 characters and entries within information columns can contain a maximum of 256 characters. Bracket characters "[" and "]" are not allowed in information column names. The specification file can contain up to 20 information columns. Same rules mentioned here applies for both "Parametric" and "Functional" type specifications.',
        ),
    ]

    for col_type, abbr, desc in column_types:
        ws[f"A{current_row}"] = col_type
        ws[f"B{current_row}"] = abbr
        ws[f"C{current_row}"] = desc
        ws[f"C{current_row}"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[current_row].height = 40
        current_row += 1

    current_row += 1

    # Standard Column Overview
    ws[f"A{current_row}"] = "Standard Column Overview"
    ws[f"A{current_row}"].font = Font(bold=True, size=11)
    current_row += 1

    ws[f"A{current_row}"] = "Column Name"
    ws[f"B{current_row}"] = "Description"
    ws[f"C{current_row}"] = "Entry Required?"
    ws[f"D{current_row}"] = "Specification Types Required In?"
    ws[f"E{current_row}"] = "Entry Format Requirements"
    for col in ["A", "B", "C", "D", "E"]:
        ws[f"{col}{current_row}"].font = Font(bold=True)
        ws[f"{col}{current_row}"].fill = description_fill
    current_row += 1

    standard_columns = [
        (
            "SpecID",
            "ID for product specification defined by user.",
            "Yes",
            "Both Parametric and Functional",
            "Unique entries per file. Maximum limit of 32 characters for an entry.",
        ),
        (
            "Block",
            "IP block for the product specification.",
            "No",
            "Both Parametric and Functional",
            "",
        ),
        (
            "Spec Symbol",
            "Shorthand identifier for the product specification.",
            "No",
            "Both Parametric and Functional",
            "",
        ),
        (
            "Spec Name",
            "Descriptive name of the product specification.",
            "Yes",
            "Both Parametric and Functional",
            "",
        ),
        (
            "Min",
            "Minimum acceptable value for the product specification during testing.",
            "See Notes*",
            "Only Parametric",
            "Numerical values only. Entry value must be less than Max column entry.",
        ),
        (
            "Typical",
            "Typical value for the product specification during testing.",
            "See Notes*",
            "Only Parametric",
            "Numerical values only.",
        ),
        (
            "Max",
            "Maximum acceptable value for the product specification during testing.",
            "See Notes*",
            "Only Parametric",
            "Numerical values only. Entry value must be greater than Min column entry.",
        ),
        ("Unit", "Unit of measure for the product specification.", "No", "Only Parametric", ""),
    ]

    for name, desc, required, spec_types, format_req in standard_columns:
        ws[f"A{current_row}"] = name
        ws[f"B{current_row}"] = desc
        ws[f"C{current_row}"] = required
        ws[f"D{current_row}"] = spec_types
        ws[f"E{current_row}"] = format_req
        for col in ["B", "E"]:
            ws[f"{col}{current_row}"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[current_row].height = 15
        current_row += 1

    current_row += 1

    # Standard Column Notes
    ws[f"A{current_row}"] = "Standard Column Notes"
    ws[f"A{current_row}"].font = Font(bold=True, size=11)
    current_row += 1
    ws[f"A{current_row}"] = (
        "* Each product specification must have a value set for at least one of the Min, "
        "Max, or Typical columns."
    )
    ws[f"A{current_row}"].alignment = Alignment(wrap_text=True)
    current_row += 2

    # Condition Column Formatting
    ws[f"A{current_row}"] = "Condition Column Formatting"
    ws[f"A{current_row}"].font = Font(bold=True, size=11)
    current_row += 1
    ws[f"A{current_row}"] = (
        'Data in condition columns must be enclosed in square brackets, "[" and "]", '
        "and must adhere to one of the following formats."
    )
    ws[f"A{current_row}"].alignment = Alignment(wrap_text=True)
    current_row += 1

    ws[f"A{current_row}"] = "Format"
    ws[f"B{current_row}"] = "Examples"
    for col in ["A", "B"]:
        ws[f"{col}{current_row}"].font = Font(bold=True)
        ws[f"{col}{current_row}"].fill = description_fill
    current_row += 1

    condition_formats = [
        (
            "Comma-separated numerical data. Space between value and comma is optional.",
            "[3]\n[2,-3,5]\n[31, 2E+6, 36]",
        ),
        (
            'Single-sided range of numeric values separated by ".." to indicate that the values are either:\nLess than or equal to X, represented by ".." before the value.\nGreater than or equal to X, represented by ".." after the value.',
            "[-3..]\n[-3e2..5..9..]\n[..11]\n[..11e-3..15E-2..19]",
        ),
        (
            'Double-sided range of numeric values separated by ".." to indicate a range of values between two numbers.',
            "[3..16]\n[-13..9..31]\n[3..12..24..36]",
        ),
    ]

    for format_desc, examples in condition_formats:
        ws[f"A{current_row}"] = format_desc
        ws[f"B{current_row}"] = examples
        ws[f"A{current_row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"B{current_row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[current_row].height = 40
        current_row += 1

    current_row += 1

    # Condition Value Notes
    ws[f"A{current_row}"] = "Condition Value Notes"
    ws[f"A{current_row}"].font = Font(bold=True, size=11)
    current_row += 1
    ws[f"A{current_row}"] = (
        "The total number of condition values for a spec should be less than 150."
    )
    current_row += 1

    ws[f"A{current_row}"] = "Note"
    ws[f"B{current_row}"] = "Example(s)"
    for col in ["A", "B"]:
        ws[f"{col}{current_row}"].font = Font(bold=True)
        ws[f"{col}{current_row}"].fill = description_fill
    current_row += 1

    condition_value_notes = [
        (
            "Range is considered as single condition value.",
            "[0..2..5]\ncondition value = 1\n[..3..5]\ncondition value = 1\n[2..5..]\ncondition value = 1",
        ),
        (
            "List with n comma separated values have n condition values",
            "[1,2,3,4,5]\ncondition value = 5\n[-2,-1,0]\ncondition value = 3\n[-15]\ncondition value = 1",
        ),
    ]

    for note, examples in condition_value_notes:
        ws[f"A{current_row}"] = note
        ws[f"B{current_row}"] = examples
        ws[f"A{current_row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"B{current_row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[current_row].height = 40
        current_row += 1

    current_row += 1

    # Condition Value Range Notes
    ws[f"A{current_row}"] = "Condition Value Range Notes"
    ws[f"A{current_row}"].font = Font(bold=True, size=11)
    current_row += 1

    ws[f"A{current_row}"] = "Note"
    ws[f"B{current_row}"] = "Example(s)"
    for col in ["A", "B"]:
        ws[f"{col}{current_row}"].font = Font(bold=True)
        ws[f"{col}{current_row}"].fill = description_fill
    current_row += 1

    condition_range_notes = [
        (
            "Values must be in ascending order left-to-right.",
            "[31..56..101]\n[-25..-10..0..23]\n[3..6..91..]\n[..-6..0..15..30]",
        ),
        (
            "Values specified in the range are required condition values. Data points for these values must be present in the product specification data to indicate a passing condition.",
            "[-3..15..33..96]\n-3, 15, 33, and 96 are required values for this condition.",
        ),
        (
            "The first and last values in a range are inclusive. In double-sided ranges, the first value indicates the minimum and the last value indicates the maximum extent of the range.",
            "[-3..15..33..96]\n-3 and 96 are included in the range, -3 <= x <= 96",
        ),
        (
            "In single-sided ranges, only the minimum or maximum is provided.",
            "[3E-13..6..9..]\nThe minimum of the range is 3, there is no set maximum value.\n[..-6..9..12e1]\nThe maximum value of the range is 12, there is no set minimum value.",
        ),
    ]

    for note, examples in condition_range_notes:
        ws[f"A{current_row}"] = note
        ws[f"B{current_row}"] = examples
        ws[f"A{current_row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"B{current_row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[current_row].height = 40
        current_row += 1

    # Adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 40

    # Create Parametric example sheet
    ws_param = wb.create_sheet("Spec_Parametric Example")
    ws_param["A1"] = "Category Name"
    ws_param["B1"] = "Power Supply"
    ws_param["A2"] = "Type"
    ws_param["B2"] = "Parametric"

    # Row 3: Column types
    # Row 3: Column types (aligned with headers in row 4)
    column_types_row = [
        "STD",
        "STD",
        "STD",
        "STD",
        "STD",
        "STD",
        "STD",
        "STD",
        "COND",
        "COND",
        "INF",
    ]
    for idx, col_type in enumerate(column_types_row, 1):
        ws_param.cell(row=3, column=idx).value = col_type
        ws_param.cell(row=3, column=idx).fill = column_type_fill
        ws_param.cell(row=3, column=idx).font = Font(bold=True)

    # Row 4: Column headers (must match SPEC_FILE_MAPPING keys for STD columns)
    headers = [
        "Spec ID",
        "Block",
        "Spec Symbol",
        "Spec Name",
        "Min",
        "Typical",
        "Max",
        "Unit",
        "VCC (V)",
        "Temp (°C)",
        "Notes",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_param.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = std_column_fill

    # Sample spec data
    sample_specs = [
        (
            "1",
            "Amplifier",
            "Vin",
            "Input Voltage",
            "10",
            "12",
            "14",
            "V",
            "[5,12]",
            "[25,85]",
            "Standard input",
        ),
        (
            "2",
            "Amplifier",
            "Vout",
            "Output Voltage",
            "1",
            "5",
            "9",
            "V",
            "[5]",
            "[0,50]",
            "Primary output",
        ),
        (
            "3",
            "Amplifier",
            "Ib",
            "Input Bias Current",
            "10",
            "50",
            "100",
            "nA",
            "[5,12]",
            "[25]",
            "Typical at 25C",
        ),
    ]

    for row_idx, spec in enumerate(sample_specs, 5):
        for col_idx, value in enumerate(spec, 1):
            cell = ws_param.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Adjust column widths
    param_widths = [10, 12, 12, 20, 10, 10, 10, 8, 12, 12, 20]
    for col_idx, width in enumerate(param_widths, 1):
        ws_param.column_dimensions[get_column_letter(col_idx)].width = width

    # Create Functional example sheet
    ws_func = wb.create_sheet("Spec_Functional Example")
    ws_func["A1"] = "Category Name"
    ws_func["B1"] = "Thermal Characteristics"
    ws_func["A2"] = "Type"
    ws_func["B2"] = "Functional"

    # Row 3: Column types (aligned with headers in row 4)
    column_types_func = ["STD", "STD", "STD", "STD", "COND", "COND", "INF"]
    for idx, col_type in enumerate(column_types_func, 1):
        ws_func.cell(row=3, column=idx).value = col_type
        ws_func.cell(row=3, column=idx).fill = column_type_fill
        ws_func.cell(row=3, column=idx).font = Font(bold=True)

    # Row 4: Column headers (must match SPEC_FILE_MAPPING keys for STD columns)
    headers_func = [
        "Spec ID",
        "Block",
        "Spec Symbol",
        "Spec Name",
        "Ambient Temp (°C)",
        "Power (mW)",
        "Notes",
    ]
    for col_idx, header in enumerate(headers_func, 1):
        cell = ws_func.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = std_column_fill

    # Sample functional specs
    sample_func_specs = [
        (
            "1",
            "Thermal",
            "Tj_max",
            "Max Junction Temperature",
            "[0..125]",
            "[1,10]",
            "Operating range",
        ),
        ("2", "Thermal", "Theta_ja", "Junction to Ambient", "[50..100]", "[10]", "Still air"),
    ]

    for row_idx, spec in enumerate(sample_func_specs, 5):
        for col_idx, value in enumerate(spec, 1):
            cell = ws_func.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Adjust column widths
    func_widths = [10, 12, 12, 20, 15, 15, 20]
    for col_idx, width in enumerate(func_widths, 1):
        ws_func.column_dimensions[get_column_letter(col_idx)].width = width

    # Save file
    output_path = (
        Path(__file__).parent.parent / "slcli/examples/spec-compliance-notebooks/spec_template.xlsx"
    )
    wb.save(output_path)
    print(f"✓ Spec template created: {output_path}")


if __name__ == "__main__":
    create_spec_template()
